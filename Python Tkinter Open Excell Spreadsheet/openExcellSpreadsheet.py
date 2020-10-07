# Python Tkinter Open Excell Spreadsheet
# Hoja de cálculo de Python Tkinter Open Excell

from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title('Python Tkinter Open Excell Spreadsheetx')
root.iconbitmap('Python Tkinter Open Excell Spreadsheet/icons/excel.ico')
root.geometry("500x500")

#my_list =["one", "two", "three"]

# select function
def select(e):
    my_label.config(text=my_listbox.get(ANCHOR))



# Create ListBox
my_listbox = Listbox(root, width=45)
my_listbox.pack(pady=20)

#for item in my_list:
#    my_listbox.insert(END, item)

# Create a wb
wb = load_workbook('Python Tkinter Open Excell Spreadsheet/documents/name_color.xlsx')
# set active worksheet
ws = wb.active

# Grab a column of data
col_a = ws["A"]
col_b = ws["B"]

for item in col_a:
    my_listbox.insert(END, item.value)

my_label = Label(root, text="Select Item...", font=("Helvetica", 18))
my_label.pack(pady=20)

# Create listbox binding
my_listbox.bind("<ButtonRelease-1>", select)

root.mainloop()
