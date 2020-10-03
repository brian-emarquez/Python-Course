#----------------------------------------------------------#
# Python Tkinter Using Excel Spreadsheets
# Python Tkinter usando hojas de cálculo de Excel

# pip install openpyxl

#----------------------------------------------------------#

from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()

root.title('Python Tkinter Using Excel Spreadsheets')
root.iconbitmap('Python Tkinter Using Excel Spreadsheets/Icons/file.ico')
root.geometry("500x650")

# Create workbook instance
wb = Workbook()
# Load existing workbook
wb = load_workbook('Python Tkinter Using Excel Spreadsheets/documents/pizza.xlsx')
# Create active worksheet
ws = wb.active

# Create variable for Column A
column_a = ws ['A']
column_b = ws ['B']

def get_a():
    list= " "
    for cell in column_a:
        list = f'{list + str(cell.value)}\n'

    label_a.config(text=list)


def get_b():
    list= " "
    for cell in column_b:
        list = f'{list + str(cell.value)}\n'

    label_b.config(text=list)


ba = Button(root, text="Get Column A", command=get_a)
ba.pack(pady=20)

label_a = Label(root, text="")
label_a.pack(pady=20)

bb = Button(root, text="Get Column B", command=get_b)
bb.pack(pady=20)

label_b = Label(root, text="")
label_b.pack(pady=20)

print(column_a)

# add fred to A8 and B9
ws["A8"]  ="Fred"
ws["B8"]  ="Cheese"

# Save new File
wb.save('Python Tkinter Using Excel Spreadsheets/documents/pizza2.xlsx')


root.mainloop()
